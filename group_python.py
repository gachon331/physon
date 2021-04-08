# -*- coding: utf-8 -*-
"""
Created on Thu Apr  8 22:41:07 2021

@author: ng110
"""

from openpyxl import Workbook
wb = Workbook()
ws = wb.active 
ws['A1'] = 'st_num'
ws['B1'] = 'name'
ws['C1'] = 'group_id'
ws['D1'] = 'git'
ws.title = 'jo1'

ws2 = wb.create_sheet(1, 'jo2') 
ws2['A1'] = 'st_num'
ws2['B1'] = 'name'
ws2['C1'] = 'group_id'
ws2['D1'] = 'git'

ws3 = wb.create_sheet(1, 'jo3') 
ws3['A1'] = 'st_num'
ws3['B1'] = 'name'
ws3['C1'] = 'group_id'
ws3['D1'] = 'git'

ws4 = wb.create_sheet(1, 'jo4') 
ws4['A1'] = 'st_num'
ws4['B1'] = 'name'
ws4['C1'] = 'group_id'
ws4['D1'] = 'git'

ws5 = wb.create_sheet(1, 'jo5') 
ws5['A1'] = 'st_num'
ws5['B1'] = 'name'
ws5['C1'] = 'group_id'
ws5['D1'] = 'git'

ws6 = wb.create_sheet(1, 'jo6') 
ws6['A1'] = 'st_num'
ws6['B1'] = 'name'
ws6['C1'] = 'group_id'
ws6['D1'] = 'git'

ws7 = wb.create_sheet(1, 'jo7') 
ws7['A1'] = 'st_num'
ws7['B1'] = 'name'
ws7['C1'] = 'group_id'
ws7['D1'] = 'git'

ws8 = wb.create_sheet(1, 'jo8') 
ws8['A1'] = 'st_num'
ws8['B1'] = 'name'
ws8['C1'] = 'group_id'
ws8['D1'] = 'git'

ws9 = wb.create_sheet(1, 'jo9') 
ws9['A1'] = 'st_num'
ws9['B1'] = 'name'
ws9['C1'] = 'group_id'
ws9['D1'] = 'git'

ws10 = wb.create_sheet(1, 'jo10') 
ws10['A1'] = 'st_num'
ws10['B1'] = 'name'
ws10['C1'] = 'group_id'
ws10['D1'] = 'git'

wb.save('group_python.xlsx')